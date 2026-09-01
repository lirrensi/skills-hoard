"""Excel Formula Recalculation.

Usage:
    python scripts/recalc.py <excel_file> [timeout_seconds]

Returns JSON with error details:
    {"status": "success", "total_errors": 0, "total_formulas": 42}
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALC_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def get_env():
    env = os.environ.copy()
    if sys.platform == "win32":
        for p in [
            r"C:\Program Files\LibreOffice\program",
            r"C:\Program Files (x86)\LibreOffice\program",
        ]:
            if Path(p).exists():
                env["PATH"] = p + os.pathsep + env.get("PATH", "")
                break
    return env


def setup_macro():
    macro_dir = os.path.expanduser(
        MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX
    )
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text()
    ):
        return True

    try:
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=get_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)
        Path(macro_file).write_text(RECALC_MACRO)
        return True
    except Exception:
        return False


def recalc(filename: str, timeout: int = 30) -> dict:
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    if not setup_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        str(Path(filename).absolute()),
    ]

    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_env())

    if result.returncode != 0 and result.returncode != 124:
        return {"error": result.stderr or "Recalculation failed"}

    try:
        wb = load_workbook(filename, data_only=True)
        errors = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}
        error_details = {e: [] for e in errors}
        total = 0

        for sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        for err in errors:
                            if err in cell.value:
                                error_details[err].append(f"{sheet}!{cell.coordinate}")
                                total += 1
                                break
        wb.close()

        result = {
            "status": "success" if total == 0 else "errors_found",
            "total_errors": total,
        }

        for err_type, locs in error_details.items():
            if locs:
                result["error_summary"] = result.get("error_summary", {})
                result["error_summary"][err_type] = {
                    "count": len(locs),
                    "locations": locs[:20],
                }

        wb2 = load_workbook(filename, data_only=False)
        formula_count = sum(
            1
            for s in wb2.sheetnames
            for r in wb2[s].iter_rows()
            for c in r
            if c.value and isinstance(c.value, str) and c.value.startswith("=")
        )
        wb2.close()
        result["total_formulas"] = formula_count

        return result
    except Exception as e:
        return {"error": str(e)}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    print(json.dumps(recalc(filename, timeout), indent=2))
